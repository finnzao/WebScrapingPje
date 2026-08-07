"""
Captura de juntadas por tipo de documento e período — orientada a fila.

Responde ao problema: dado um lote de NPUs (vindos do Datajud e do painel
gerencial), identificar quais processos tiveram juntada de um tipo de
documento dentro de uma janela de datas e baixar o PDF correspondente.
Usa o filtro tipo+data da própria tela de autos digitais, então o PDF
gerado contém APENAS os documentos da janela — não os autos inteiros.

Resultado por processo (tabela fila, ver fila.py):
  capturado    -> houve juntada no período; PDF baixado (coluna arquivo)
  sem_juntada  -> consulta OK, nenhum documento do tipo no período
  erro         -> erro_tipo diz o motivo (sigilo, timeout, sessao_expirada,
                  nao_encontrado, pdf_invalido, resposta_desconhecida...)

Uso:
  python capturarJuntadasPorPeriodo.py --tudo --dias 16 --perfil "Direção de Secretaria"
  python capturarJuntadasPorPeriodo.py --importar processos.csv --origem datajud
  python capturarJuntadasPorPeriodo.py --importar-painel --dias 15
  python capturarJuntadasPorPeriodo.py --importar-painel --ignorar-tarefa "Acompanhar MPU"
  python capturarJuntadasPorPeriodo.py --executar
  python capturarJuntadasPorPeriodo.py --executar --tipo Peticao --inicio 21/07/2026 --fim 01/08/2026
  python capturarJuntadasPorPeriodo.py --status
  python capturarJuntadasPorPeriodo.py --reprocessar timeout
  python capturarJuntadasPorPeriodo.py --exportar resultado.csv

Retomável: pare com Ctrl+C e rode --executar de novo; continua de onde
parou. Para concorrência, abra 2-3 terminais com --executar (o claim da
fila é atômico). Mais que isso é descortesia com o servidor do PJe.
"""

import argparse
import csv
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import fila
from pje_lib import PJEClient, SessaoExpirada, ErroCaptura
from pje_lib.services.task_service import normalizar_texto
from pje_lib.utils import delay

# Janela e tipo padrão do recorte atual (Rio Real/BA)
TIPO_PADRAO = "Peticao"
INICIO_PADRAO = "21/07/2026"
FIM_PADRAO = "01/08/2026"

DOWNLOAD_DIR = fila.DATA_DIR / "capturas"
# Quantas solicitações enviadas à área de download antes de parar para coletá-las
TAMANHO_LOTE_AREA = 10


def normalizar_npu(bruto: str) -> str:
    """Aceita NPU com ou sem pontuação e devolve o formato CNJ
    (NNNNNNN-DD.AAAA.J.TR.OOOO), que é como o PJE exibe e pesquisa."""
    digitos = re.sub(r"\D", "", str(bruto))
    if len(digitos) != 20:
        return str(bruto).strip()
    return (f"{digitos[:7]}-{digitos[7:9]}.{digitos[9:13]}."
            f"{digitos[13]}.{digitos[14:16]}.{digitos[16:]}")


def ler_arquivo(caminho: Path, origem_padrao: str):
    """Lê .csv (npu[;origem]) ou .xlsx (1ª coluna npu, 2ª origem opcional)."""
    linhas = []
    if caminho.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook  # já presente no ambiente de planilhas
        ws = load_workbook(caminho, read_only=True).active
        iterador = ([str(c) if c is not None else "" for c in row]
                    for row in ws.iter_rows(values_only=True))
    else:
        texto = caminho.read_text(encoding="utf-8-sig").splitlines()
        if not texto:
            return []
        iterador = csv.reader(texto, delimiter=";" if ";" in texto[0] else ",")
    for row in iterador:
        if not row or not str(row[0]).strip():
            continue
        npu = normalizar_npu(row[0])
        if not re.search(r"\d{7}-?\d{2}", npu):
            continue  # cabeçalho ou lixo
        origem = (row[1].strip().lower() if len(row) > 1 and str(row[1]).strip() in ("datajud", "painel")
                  else origem_padrao)
        linhas.append((npu, origem))
    return linhas


def coletar_area_download(pje: PJEClient, con, pendentes_area: list):
    """Aguarda e baixa o que foi enviado para a área de download,
    atualizando a fila por processo."""
    if not pendentes_area:
        return
    downloads = pje._downloads.aguardar_downloads(pendentes_area, tempo_maximo=300)
    baixados = set()
    for dl in downloads:
        delay()
        try:
            arquivo = pje._downloads.baixar_arquivo(dl, DOWNLOAD_DIR)
        except Exception as e:
            arquivo = None
            for npu in dl.get_numeros_processos():
                if npu in pendentes_area:
                    fila.falhar(con, npu, "pdf_invalido", str(e))
                    baixados.add(npu)
        if arquivo:
            for npu in dl.get_numeros_processos():
                if npu in pendentes_area:
                    fila.concluir(con, npu, arquivo=str(arquivo))
                    baixados.add(npu)
    for npu in pendentes_area:
        if npu not in baixados:
            fila.falhar(con, npu, "nao_chegou_area_download",
                        "solicitado mas não apareceu na área de download")
    pendentes_area.clear()


def criar_cliente(args) -> PJEClient:
    pje = PJEClient(download_dir=str(DOWNLOAD_DIR),
                    session_dir=str(fila.DATA_DIR / "session"),
                    log_dir=str(fila.DATA_DIR / "logs"))
    if not pje.login():
        print("Login falhou. Verifique o .env")
        sys.exit(1)
    if args.perfil:
        if not pje.select_profile(args.perfil):
            print(f"Perfil '{args.perfil}' não encontrado. Veja os nomes com --listar-perfis")
            sys.exit(1)
    else:
        # Sem --perfil roda com o perfil da sessão atual — deixa isso explícito,
        # porque o universo de tarefas/processos muda conforme o perfil
        print("[aviso] Rodando com o perfil atual da sessão (use --perfil para trocar, "
              "--listar-perfis para ver as opções)")
    return pje


def listar_perfis(args):
    pje = PJEClient(download_dir=str(DOWNLOAD_DIR),
                    session_dir=str(fila.DATA_DIR / "session"),
                    log_dir=str(fila.DATA_DIR / "logs"))
    if not pje.login():
        print("Login falhou. Verifique o .env")
        sys.exit(1)
    try:
        perfis = pje.listar_perfis()
        if not perfis:
            print("Nenhum perfil encontrado")
            return
        print(f"\n{len(perfis)} perfis disponíveis (use o nome com --perfil):\n")
        for p in perfis:
            print(f"  - {p.nome_completo}")
    finally:
        pje.close()


# Radical que casa com as variações de tarefa de arquivo do acervo:
# "(CR) Arquivado - Provisório", "(CR) Arquivar processo", "Arquivo definitivo",
# "(EF) Arquivado - Definitivo", "Arquivado provisoriamente", "Arquivo provisorio"...
PADRAO_ARQUIVADAS = "arquiv"


def tarefa_ignorada(nome: str, ignorar_extra: list, incluir_arquivados: bool) -> bool:
    """Decide se uma tarefa fica fora da descoberta (comparação sem acento/caixa)."""
    nome_norm = normalizar_texto(nome)
    if not incluir_arquivados and PADRAO_ARQUIVADAS in nome_norm:
        return True
    return any(normalizar_texto(pat) in nome_norm for pat in ignorar_extra)


def importar_painel(args):
    """Descoberta direta do PJE: varre as tarefas do perfil atual e alimenta
    a fila com origem='painel' — mesma mecânica da extensão PJe R+
    (endpoints do painel + filtro no cliente), sem planilha intermediária."""
    con = fila.conectar()
    pje = criar_cliente(args)
    ignorar_extra = args.ignorar_tarefa or []
    corte_ms = (time.time() - args.dias * 86400) * 1000 if args.dias else None

    processos = {}  # npu -> id_processo
    sem_data = 0
    avaliados = 0
    try:
        tarefas = pje.listar_tarefas()
        print(f"{len(tarefas)} tarefas no perfil atual")
        for t in tarefas:
            if tarefa_ignorada(t.nome, ignorar_extra, args.incluir_arquivados):
                print(f"  [ignorada] {t.nome}")
                continue
            no_filtro = 0
            for p in pje.listar_processos_tarefa(t.nome):
                avaliados += 1
                # filtro por último movimento: juntada não tira o processo da
                # tarefa, então dataChegada sozinha deixaria passar batido
                if corte_ms and p.data_referencia and p.data_referencia < corte_ms:
                    continue
                if corte_ms and not p.data_referencia:
                    sem_data += 1  # sem info de movimento: mantém, não descarta às cegas
                # guarda o idProcesso do payload: a captura usa direto,
                # sem precisar resolver o NPU via consulta
                processos.setdefault(normalizar_npu(p.numero_processo), p.id_processo or None)
                no_filtro += 1
            print(f"  [lendo] {t.nome}: {no_filtro}/{t.quantidade_pendente} no filtro")
            delay(0.5, 1.5)
    finally:
        pje.close()

    fila.importar(con, [(n, "painel", idp) for n, idp in sorted(processos.items())])
    filtro_txt = f" com movimento nos últimos {args.dias} dias" if args.dias else ""
    print(f"\n{avaliados} pendências avaliadas -> {len(processos)} processos únicos{filtro_txt}")
    if sem_data:
        print(f"(aviso: {sem_data} sem ultimoMovimento/dataChegada no payload — incluídos apesar do --dias)")
    print("Fila:", fila.resumo(con))


def relatorio_final(con, args):
    """Fecha a análise: o que foi avaliado, o que foi encontrado, onde está."""
    total = con.execute("SELECT COUNT(*) FROM fila").fetchone()[0]
    capturados = con.execute(
        "SELECT npu, arquivo FROM fila WHERE status='capturado' ORDER BY npu").fetchall()
    sem_juntada = con.execute(
        "SELECT COUNT(*) FROM fila WHERE status='sem_juntada'").fetchone()[0]
    pendentes = con.execute(
        "SELECT COUNT(*) FROM fila WHERE status='pendente'").fetchone()[0]
    erros = con.execute(
        """SELECT erro_tipo, COUNT(*) FROM fila WHERE status='erro'
           GROUP BY erro_tipo ORDER BY COUNT(*) DESC""").fetchall()

    print("\n" + "=" * 62)
    print("RELATÓRIO DA ANÁLISE")
    print("=" * 62)
    print(f"Filtro: documentos '{args.tipo}' juntados entre {args.inicio} e {args.fim}")
    print(f"Processos na fila: {total}  (analisados: {total - pendentes}, pendentes: {pendentes})")
    print(f"\nCOM JUNTADA no período: {len(capturados)} processo(s)")
    for npu, arquivo in capturados[:50]:
        n_docs = len(arquivo.split(";")) if arquivo else 0
        print(f"  {npu}  ({n_docs} PDF{'s' if n_docs != 1 else ''})")
    if len(capturados) > 50:
        print(f"  ... e mais {len(capturados) - 50} (lista completa no CSV)")
    print(f"\nSEM juntada do tipo no período: {sem_juntada} processo(s)")
    if erros:
        print("\nErros (reprocessáveis com --reprocessar <tipo>):")
        for tipo, n in erros:
            print(f"  {tipo}: {n}")
    print(f"\nPDFs em: {DOWNLOAD_DIR}")

    saida = Path.cwd() / f"resultado_captura_{datetime.now():%Y%m%d_%H%M}.csv"
    exportar(con, saida)


def executar(args):
    con = fila.conectar()
    orfaos = fila.recuperar_orfaos(con)
    if orfaos:
        print(f"[fila] {orfaos} processo(s) órfão(s) devolvidos à fila")

    pje = criar_cliente(args)

    pendentes_area = []
    processados = 0
    try:
        while not (args.limite and processados >= args.limite) and (item := fila.pegar_proximo(con)):
            npu, origem, id_processo = item
            processados += 1
            print(f"[{processados}] {npu} ({origem})")

            try:
                det = pje.capturar_documentos(
                    npu, tipo_documento=args.tipo,
                    data_inicio=args.inicio, data_fim=args.fim,
                    diretorio=DOWNLOAD_DIR,
                    id_processo=id_processo,
                )
                if det.get("sucesso"):
                    if det.get("arquivo_baixado"):
                        fila.concluir(con, npu, arquivo=det["arquivo_baixado"])
                        print(f"    capturado (direto): {det['arquivo_baixado']}")
                    else:
                        pendentes_area.append(npu)
                        print("    enviado para área de download")
                elif det.get("erro_tipo") == "sem_documentos":
                    fila.sem_juntada(con, npu)
                    print(f"    sem juntada de '{args.tipo}' no período")
                else:
                    motivo = det.get("erro_msg") or str(det.get("mensagens", ""))
                    fila.falhar(con, npu, det.get("erro_tipo", "desconhecido"), motivo)
                    print(f"    erro: {det.get('erro_tipo')} — {motivo[:120]}")
            except SessaoExpirada as e:
                # _com_relogin já tentou religar uma vez; aqui é falha real
                fila.falhar(con, npu, e.erro_tipo, str(e))
                print(f"    erro: sessão expirada e religação falhou")
            except ErroCaptura as e:
                fila.falhar(con, npu, e.erro_tipo, str(e))
                print(f"    erro: {e.erro_tipo} — {e}")
            except Exception as e:
                fila.falhar(con, npu, "excecao", f"{type(e).__name__}: {e}")
                print(f"    erro inesperado: {e}")

            if len(pendentes_area) >= TAMANHO_LOTE_AREA:
                coletar_area_download(pje, con, pendentes_area)

            delay(2, 4)  # cortesia com o servidor
    except KeyboardInterrupt:
        print("\nInterrompido. Estado salvo — rode --executar para retomar.")
    finally:
        coletar_area_download(pje, con, pendentes_area)
        pje.close()

    relatorio_final(con, args)


def exportar(con, caminho: Path):
    linhas = con.execute(
        """SELECT npu, origem, status, via, tentativas, erro_tipo, erro_msg,
                  arquivo, atualizado_em
           FROM fila ORDER BY status, npu""").fetchall()
    with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["numeroProcesso", "origem", "status", "via", "tentativas",
                    "erro_tipo", "erro_msg", "arquivo", "atualizado_em"])
        w.writerows(linhas)
    print(f"Exportado: {caminho} ({len(linhas)} linhas)")


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--importar", metavar="ARQUIVO", help="csv/xlsx com NPUs para a fila")
    p.add_argument("--origem", choices=["datajud", "painel"], default="painel",
                   help="origem dos NPUs importados (padrão: painel)")
    p.add_argument("--tudo", action="store_true",
                   help="fluxo completo: descobre no painel, captura e entrega o relatório")
    p.add_argument("--importar-painel", action="store_true",
                   help="descobre processos direto das tarefas do perfil atual (sem planilha)")
    p.add_argument("--ignorar-tarefa", action="append", metavar="NOME",
                   help="tarefa a ignorar na descoberta (repetível; casa por trecho do nome)")
    p.add_argument("--incluir-arquivados", action="store_true",
                   help="inclui tarefas de arquivo (ignoradas por padrão)")
    p.add_argument("--dias", type=int, metavar="N",
                   help="só processos que chegaram na tarefa nos últimos N dias")
    p.add_argument("--executar", action="store_true", help="processa a fila")
    p.add_argument("--tipo", default=TIPO_PADRAO, help=f"tipo de documento (padrão: {TIPO_PADRAO})")
    p.add_argument("--inicio", default=INICIO_PADRAO, help=f"dd/mm/aaaa (padrão: {INICIO_PADRAO})")
    p.add_argument("--fim", default=FIM_PADRAO, help=f"dd/mm/aaaa (padrão: {FIM_PADRAO})")
    p.add_argument("--perfil", help="perfil do PJE a usar (casa por trecho do nome)")
    p.add_argument("--listar-perfis", action="store_true",
                   help="lista os perfis disponíveis e sai")
    p.add_argument("--limite", type=int, help="máximo de processos nesta execução")
    p.add_argument("--status", action="store_true", help="resumo da fila")
    p.add_argument("--reprocessar", nargs="?", const="", metavar="ERRO_TIPO",
                   help="devolve erros à fila (opcionalmente só uma categoria)")
    p.add_argument("--exportar", metavar="ARQUIVO", help="exporta a fila para csv")
    args = p.parse_args()

    if args.importar:
        con = fila.conectar()
        linhas = ler_arquivo(Path(args.importar), args.origem)
        novos = fila.importar(con, linhas)
        print(f"{len(linhas)} NPUs lidos, {novos} novos na fila (banco: {fila.DB_PATH})")
    if args.listar_perfis:
        listar_perfis(args)
        return
    if args.tudo:
        print("=" * 62)
        print("ETAPA 1/2 — Descoberta no painel")
        print("=" * 62)
        importar_painel(args)
        print("\n" + "=" * 62)
        print("ETAPA 2/2 — Captura dos documentos")
        print("=" * 62)
        executar(args)
        return
    if args.importar_painel:
        importar_painel(args)
    if args.reprocessar is not None:
        con = fila.conectar()
        n = fila.reprocessar(con, args.reprocessar or None)
        print(f"{n} processo(s) devolvido(s) à fila")
    if args.status:
        print(fila.resumo(fila.conectar()))
    if args.exportar:
        exportar(fila.conectar(), Path(args.exportar))
    if args.executar:
        executar(args)
    if not any([args.importar, args.importar_painel, args.status,
                args.reprocessar is not None, args.exportar, args.executar,
                args.listar_perfis, args.tudo]):
        p.print_help()


if __name__ == "__main__":
    main()
